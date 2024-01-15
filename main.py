import openpyxl

class SexOffender:
    def __init__(self, name, city, facebook_link, in_red):
        self.name = name
        self.city = city
        self.facebook_link = facebook_link
        self.in_red = in_red

    def to_json(self):
        return {
            "name": self.name,
            "city": self.city,
            "facebook_link": self.facebook_link,
            "in_red": self.in_red,
        }


def is_cell_red(cell):
    return cell.font.color.rgb != "FF000000"  # Red color in RGB


# Open Excel file
workbook = openpyxl.load_workbook('input.xlsx')
sheet = workbook.active

# Read Excel file and create instances of SexOffender
offenders = []
row_number = 1
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_number += 1
    name = row[0]  # Strip leading/trailing whitespaces
    city = row[1]
    facebook_link = row[2] if len(row) > 2 else None  # In case Facebook link is missing

    # Find the cell containing the name
    name_cell = sheet.cell(row=row_number, column=1)
    if name_cell.value == '':
        name_cell = sheet.cell(row=row_number, column=2)

    in_red = is_cell_red(name_cell)

    print(name, "name")
    print(name_cell.value, "value")
    if city is not None and name is not None:
        offenders.append(SexOffender(name, city, facebook_link, in_red))

# Write instances to a new file
with open('offenders.js', 'w') as output_file:
    output_file.write("const offenders = [\n")
    for offender in offenders:
        mname = offender.name.replace('"', '\\"')
        mcity = offender.city.replace('"', '\\"')
        output_file.write(f'''{{
          "name": "{mname}",
          "city": "{mcity}",
          "facebook": "{offender.facebook_link}",
          "visited": {str(offender.in_red).lower()}
        }},''')
    output_file.write("];\n\n")
    output_file.write("module.exports = offenders")
